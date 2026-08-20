from flask import Flask, jsonify, render_template, redirect, url_for, request, flash, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import bcrypt, pdfplumber, io, re, os, secrets, yfinance as yf
from datetime import datetime as dt, timedelta
from models import db, User, MutualFund, Stock, Goal, UserProfile, NetWorthHistory, Family, FamilyMember, FamilyInvite
from insurance_centre import insurance_bp
from retirement_centre import retirement_bp
from wealth import wealth_bp
from wealth.services import WealthStatisticsService
from wealth.models import WealthAssetCategory
from insurance_centre.models import (
    InsurancePolicy, InsuranceNominee, InsuranceMember,
    InsuranceAddon, InsuranceDocument, InsuranceTimeline,
    InsuranceCategory, InsuranceType, PolicyStatus,
    PremiumFrequency, DocumentType, TimelineEvent
)

app = Flask(__name__)

def _get_or_create_secret_key():
    """
    Loads SECRET_KEY from instance/secret_key.txt, generating a new
    random one on first run if the file doesn't exist yet.

    Why this approach: the previous SECRET_KEY was a hardcoded literal
    string ("mywealthlens-dev-secret-change-in-production") that had
    been sitting in this repo's git history during every window it
    was made public — meaning anyone who saw the repo during those
    windows could forge a valid session cookie or CSRF token for ANY
    user, no password needed. That key is retired for good, not
    reused here.

    instance/ is already excluded from git (see .gitignore, added
    after the Phase H database-exposure cleanup), so a file stored
    there never gets committed — this generates itself automatically
    on first run and then persists across restarts, with no manual
    environment-variable setup required on a self-hosted single
    Windows machine.
    """
    os.makedirs(app.instance_path, exist_ok=True)
    key_path = os.path.join(app.instance_path, "secret_key.txt")
    if os.path.exists(key_path):
        with open(key_path, "r") as f:
            key = f.read().strip()
            if key:
                return key
    # First run, or file was empty/corrupted — generate a fresh one.
    key = secrets.token_hex(32)
    with open(key_path, "w") as f:
        f.write(key)
    return key

app.config["SECRET_KEY"] = _get_or_create_secret_key()
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///mywealthlens.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25MB upload limit
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["WTF_CSRF_TIME_LIMIT"] = 3600
db.init_app(app)

csrf = CSRFProtect(app)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Please log in to access MyWealthLens."

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.before_request
def refresh_session():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(minutes=30)


with app.app_context():
    db.create_all()

def safe_float(val, default=0.0):
    try:
        return float(str(val).strip())
    except (TypeError, ValueError):
        return default
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/signup', methods=['GET', 'POST'])
@csrf.exempt
@limiter.limit('10 per hour')
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not name or not email or not password:
            flash('All fields are required.', 'error')
            return render_template('signup.html')
        if password != confirm:
            flash('Passwords do not match.', 'error')
            return render_template('signup.html')
        if len(password) < 8:
            flash('Password must be at least 8 characters.', 'error')
            return render_template('signup.html')
        if User.query.filter_by(email=email).first():
            flash('An account with this email already exists.', 'error')
            return render_template('signup.html')
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        user = User(name=name, email=email, password=hashed)
        db.session.add(user)
        db.session.commit()
        login_user(user)
        flash(f'Welcome to MyWealthLens, {name}!', 'success')
        return redirect(url_for('dashboard'))
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
@csrf.exempt
@limiter.limit('5 per 15 minutes', methods=['POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and bcrypt.checkpw(password.encode('utf-8'), user.password.encode('utf-8')):
            login_user(user)
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
@csrf.exempt
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter_by(email=email).first()
        flash('If an account exists for that email, a reset link has been sent.', 'success')
        return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html')

def _save_snapshot(user_id, cat_totals, mfs, stocks, liabilities_total):
    """
    Save one net worth snapshot per day.

    Phase H: previously sourced its per-category figures from the
    legacy Asset model (now retired) and its liability figure from
    the Loan model (which had no CRUD UI and was always empty — so
    liabilities never actually reduced this total). Now sourced from
    WealthAsset (via WealthStatisticsService.category_breakdown(),
    passed in as cat_totals) and WealthLiability, the two authoritative
    Wealth tables. The NetWorthHistory table/columns themselves are
    untouched — existing historical rows remain exactly as stored
    (Section 14 of the Phase H spec: historical snapshots are
    immutable). Going forward, the existing category columns hold the
    closest equivalent under the new Wealth taxonomy:
      gold        -> Precious Metals
      realestate  -> Real Estate
      debt        -> Bank & Deposits + Investments (fixed-income-like)
      cash        -> not distinguished under the new taxonomy; kept at 0
      other       -> Vehicles + Business + Other
    """
    from datetime import date as _date
    today = _date.today()
    existing = NetWorthHistory.query.filter_by(
        user_id=user_id, snapshot_date=today).first()
    if existing:
        return
    equity = sum(m.value for m in mfs) + sum(s.value for s in stocks)
    gold   = cat_totals.get(WealthAssetCategory.PRECIOUS_METALS, 0)
    re_val = cat_totals.get(WealthAssetCategory.REAL_ESTATE, 0)
    debt   = (cat_totals.get(WealthAssetCategory.BANK_DEPOSITS, 0)
              + cat_totals.get(WealthAssetCategory.INVESTMENTS, 0))
    cash   = 0
    other  = (cat_totals.get(WealthAssetCategory.VEHICLES, 0)
              + cat_totals.get(WealthAssetCategory.BUSINESS, 0)
              + cat_totals.get(WealthAssetCategory.OTHER, 0))
    liab   = liabilities_total
    total  = equity + debt + gold + re_val + cash + other - liab
    snap   = NetWorthHistory(
        user_id=user_id, snapshot_date=today, total=total,
        equity=equity, debt=debt, gold=gold,
        realestate=re_val, cash=cash, other=other, liabilities=liab)
    db.session.add(snap)
    db.session.commit()

@app.route('/dashboard')
@login_required
def dashboard():
    # Phase H: asset totals now come exclusively from WealthAsset via
    # WealthStatisticsService — the same authoritative service the
    # Wealth Net Worth page uses — instead of the retired legacy
    # Asset model. MutualFund/Stock (CAS-imported holdings) are a
    # separate, unrelated feature and are untouched.
    mfs    = MutualFund.query.filter_by(user_id=current_user.id).all()
    stocks = Stock.query.filter_by(user_id=current_user.id).all()

    wstats = WealthStatisticsService(current_user.id)
    cat_totals = {b['category']: b['total'] for b in wstats.category_breakdown()}

    real_estate_value     = cat_totals.get(WealthAssetCategory.REAL_ESTATE, 0)
    precious_metals_value = cat_totals.get(WealthAssetCategory.PRECIOUS_METALS, 0)
    vehicles_value        = cat_totals.get(WealthAssetCategory.VEHICLES, 0)
    bank_deposits_value   = cat_totals.get(WealthAssetCategory.BANK_DEPOSITS, 0)
    investments_value     = cat_totals.get(WealthAssetCategory.INVESTMENTS, 0)
    business_value        = cat_totals.get(WealthAssetCategory.BUSINESS, 0)
    other_value            = cat_totals.get(WealthAssetCategory.OTHER, 0)

    mf_value    = sum(m.value for m in mfs)
    stock_value = sum(s.value for s in stocks)

    # Matches the old dashboard's "Total" exactly in spirit: sum of all
    # holdings, no liability subtraction here (the old dashboard never
    # subtracted liabilities from this hero figure either — see the
    # Wealth Net Worth page for the liability-adjusted figure).
    wealth_assets_total = wstats.total_assets()
    total_value  = wealth_assets_total + mf_value + stock_value
    asset_count  = wstats.asset_count() + len(mfs) + len(stocks)

    # Auto daily snapshot — now sourced from WealthAsset + WealthLiability
    if total_value > 0:
        _save_snapshot(current_user.id, cat_totals, mfs, stocks,
                       wstats.total_liabilities())

    # History for stacked area chart
    history = NetWorthHistory.query.filter_by(user_id=current_user.id)\
        .order_by(NetWorthHistory.snapshot_date).limit(365).all()
    history_data = [{
        'date':        h.snapshot_date.strftime('%d %b %Y'),
        'total':       h.total,
        'equity':      h.equity,
        'debt':        h.debt,
        'gold':        h.gold,
        'realestate':  h.realestate,
        'cash':        h.cash,
        'other':       h.other,
    } for h in history]

    return render_template('dashboard.html',
        user=current_user, total=total_value,
        real_estate=real_estate_value, precious_metals=precious_metals_value,
        vehicles=vehicles_value, bank_deposits=bank_deposits_value,
        investments=investments_value, business=business_value, other=other_value,
        mf=mf_value, stocks=stock_value, mf_count=len(mfs),
        stock_count=len(stocks), mutual_funds=mfs, stock_list=stocks,
        asset_count=asset_count, history_data=history_data)

@app.route('/assets')
@login_required
def assets():
    # Phase H: the standalone legacy Assets module has been retired.
    # Wealth Assets (/wealth/assets) is now the sole authoritative
    # Assets system. This redirect protects any existing bookmarks.
    return redirect(url_for('wealth.assets_listing'))

@app.route('/preferences')
@login_required
def preferences():
    try:
        db.session.execute(db.text('SELECT 1'))
        db_connected, db_status = True, 'Healthy'
    except Exception:
        db_connected, db_status = False, 'Error'
    try:
        doc_count = InsuranceDocument.query.filter_by(user_id=current_user.id).count()
    except Exception:
        doc_count = 'Not Available'
    system_health = {
        'db_connected': db_connected,
        'db_status': db_status,
        'privacy_mode': 'Local Only',
        'version': 'v1.0.0',
        'documents_stored': doc_count,
    }
    return render_template('preferences.html', user=current_user, system_health=system_health)

@app.route('/account')
@login_required
def account():
    return redirect(url_for('preferences'))

@app.route('/settings')
@login_required
def settings():
    return redirect(url_for('preferences'))

def extract_pdf_text(file_bytes, password=None):
    try:
        pdf_file = io.BytesIO(file_bytes)
        kwargs = {'password': password} if password else {}
        full_text = ''
        with pdfplumber.open(pdf_file, **kwargs) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + chr(10)
        return full_text if full_text.strip() else None
    except Exception as e:
        app.logger.warning('PDF extraction failed: %s', e)
        return None

def _clean_scheme_name(raw):
    """Clean up MF scheme names extracted from CAMS/KFintech PDFs."""
    name = raw.strip()
    # Remove option/plan suffixes that are PDF artifacts
    name = re.sub(r'\s*[-–]\s*(Regular|Direct)\s*[-–]\s*(Growth|IDCW|Dividend).*$', 
                  lambda m: m.group(0), name)
    # Remove trailing garbage: dates, page numbers, numeric artifacts
    name = re.sub(r'\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}.*$', '', name)
    name = re.sub(r'\s+Page\s+\d+.*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s{2,}', ' ', name)
    name = name.strip(" -|/\\.,")
    # Capitalise properly if all caps
    if name == name.upper() and len(name) > 4:
        name = name.title()
    return name[:100] if name else 'Unknown Fund'

def parse_cams_pdf(text):
    holdings = []
    folio_blocks = re.split(r'Folio' + r'\s*No\s*[:\.]', text, flags=re.IGNORECASE)
    for block in folio_blocks[1:]:
        lines = [l.strip() for l in block.strip().split(chr(10)) if l.strip()]
        if not lines:
            continue
        folio = lines[0].strip()
        amc   = lines[1].strip() if len(lines) > 1 else ''
        raw_scheme = lines[2].strip() if len(lines) > 2 else ''
        scheme = _clean_scheme_name(raw_scheme)
        if not folio or len(folio) > 30:
            continue
        units = nav = value = None
        for line in lines:
            u = re.search(r'Units' + r'[:\s]+([\d,]+\.?\d*)', line, re.IGNORECASE)
            n = re.search(r'NAV' + r'[^:]*:\s*([\d,]+\.?\d*)', line, re.IGNORECASE)
            v = re.search(r'Value' + r'[:\s]+([\d,]+\.?\d*)', line, re.IGNORECASE)
            if u and not units:
                try: units = float(u.group(1).replace(',', ''))
                except: pass
            if n and not nav:
                try: nav = float(n.group(1).replace(',', ''))
                except: pass
            if v and not value:
                try: value = float(v.group(1).replace(',', ''))
                except: pass
        if units and units > 0 and scheme:
            holdings.append({'folio': folio, 'amc': amc, 'scheme': scheme,
                'units': units, 'nav': nav or 0,
                'value': value or (round(units * nav, 2) if nav else 0)})
    return holdings

def _clean_stock_name(raw_name):
    """Remove PDF artifacts, page numbers, dates and junk from holding names."""
    name = raw_name.strip()
    # Remove trailing numbers, dates, balance figures
    name = re.sub(r'\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}.*$', '', name)  # dates
    name = re.sub(r'\s+\d{1,3}(?:,\d{3})*(?:\.\d+)?\s*$', '', name)  # trailing numbers
    name = re.sub(r'\s+[A-Z]{2}\d+.*$', '', name)  # ISIN-like artifacts
    name = re.sub(r'\s{2,}', ' ', name)  # collapse spaces
    # Remove common PDF noise words at end
    noise = ['DEMAT', 'NSDL', 'CDSL', 'DP ID', 'CLIENT ID', 'CIN', 'PAN', 'ISIN']
    for word in noise:
        name = re.sub(rf'\s+{word}.*$', '', name, flags=re.IGNORECASE)
    name = name.strip(" -|/\\.,")
    return name[:80] if name else 'Unknown'

def parse_cdsl_pdf(text):
    holdings = []
    if 'STATEMENT OF HOLDINGS' in text:
        start = text.find('STATEMENT OF HOLDINGS')
        text = text[start:]
    isin_re = re.compile(r'(IN[A-Z0-9]{10})')
    # Match integers AND decimals
    int_re  = re.compile(r'\b(\d+)\b')
    dec_re  = re.compile(r'(\d+\.\d+)')
    lines_list = [l.strip() for l in text.split(chr(10)) if l.strip()]
    isin_positions = []
    for idx, line in enumerate(lines_list):
        m = isin_re.search(line)
        if m:
            isin_positions.append((idx, m.group(1), line))
    for pos_idx, (idx, isin, line) in enumerate(isin_positions):
        if pos_idx + 1 < len(isin_positions):
            next_idx = isin_positions[pos_idx + 1][0]
        else:
            next_idx = min(idx + 8, len(lines_list))
        block_lines = lines_list[idx:next_idx]
        block = ' '.join(block_lines)

        # ── Fix Bug 2: SGB quantity ──
        # SGBs are identified by ISIN starting with IN0 or name containing SGB/GOLD BOND
        is_sgb = ('SGB' in block.upper() or 'GOLD BOND' in block.upper()
                  or 'SOVEREIGN' in block.upper())

        # Get all decimal numbers (quantity and value are decimals in CDSL)
        dec_nums = [float(n) for n in dec_re.findall(block)]
        # Get all integers (SGB units are integers)
        int_nums = [int(n) for n in int_re.findall(block)
                    if 0 < int(n) < 100000]

        if not dec_nums:
            continue

        value = max(dec_nums)

        if is_sgb:
            # For SGBs: quantity is the number of bonds (integer, usually small like 1,2,3)
            # Pick the smallest positive integer that makes sense as unit count
            sgb_qty_candidates = [n for n in int_nums if 0 < n <= 1000]
            quantity = float(min(sgb_qty_candidates)) if sgb_qty_candidates else 1.0
        else:
            # Normal stocks: quantity is decimal, smaller than value
            qty_candidates = [n for n in dec_nums if 0 < n < value]
            if not qty_candidates:
                continue
            quantity = qty_candidates[0]

        # ── Fix Bug 1: Clean stock/MF name ──
        raw_name = line.replace(isin, '').strip()
        name = _clean_stock_name(raw_name)

        price = round(value / quantity, 2) if quantity > 0 else 0
        holdings.append({'isin': isin, 'name': name,
            'quantity': quantity, 'price': price, 'value': value})

    # Deduplicate by ISIN
    seen = set()
    unique = []
    for h in holdings:
        if h['isin'] not in seen:
            seen.add(h['isin'])
            unique.append(h)
    return unique

def fetch_live_price_by_isin(isin, name):
    try:
        t = yf.Ticker(isin)
        price = float(t.fast_info.last_price)
        if price and price > 0:
            return isin, round(price, 2)
    except:
        pass
    try:
        ticker_guess = re.sub(r'[^A-Z0-9]', '', name.upper()[:10]) + '.NS'
        t = yf.Ticker(ticker_guess)
        price = float(t.fast_info.last_price)
        if price and price > 0:
            return ticker_guess, round(price, 2)
    except:
        pass
    return None, None
@app.route('/upload')
@login_required
def upload():
    mf_count = MutualFund.query.filter_by(user_id=current_user.id).count()
    st_count = Stock.query.filter_by(user_id=current_user.id).count()
    return render_template('upload.html', mf_count=mf_count, st_count=st_count)

@app.route('/upload/cams', methods=['POST'])
@login_required
def upload_cams():
    pdf_file = request.files.get('pdf_file')
    password = request.form.get('password', '').strip().upper()
    if not pdf_file or pdf_file.filename == '':
        flash('Please select a PDF file.', 'error')
        return redirect(url_for('upload'))
    try:
        file_bytes = pdf_file.read()
    except Exception:
        flash('Could not read the uploaded file.', 'error')
        return redirect(url_for('upload'))
    text = extract_pdf_text(file_bytes, password if password else None)
    if not text:
        flash('Could not read the PDF. Please check the password and try again.', 'error')
        return redirect(url_for('upload'))
    holdings = parse_cams_pdf(text)
    if not holdings:
        flash('No mutual fund holdings found in this PDF.', 'error')
        return redirect(url_for('upload'))
    MutualFund.query.filter_by(user_id=current_user.id).delete()
    for h in holdings:
        mf = MutualFund(user_id=current_user.id, folio=h['folio'], amc=h['amc'],
            scheme=h['scheme'], units=h['units'], nav=h['nav'],
            value=h['value'], source='cams')
        db.session.add(mf)
    db.session.commit()
    flash(f'Successfully imported {len(holdings)} mutual fund holdings!', 'success')
    return redirect(url_for('upload'))

@app.route('/upload/cdsl', methods=['POST'])
@login_required
def upload_cdsl():
    pdf_file = request.files.get('pdf_file')
    password = request.form.get('password', '').strip()
    if not pdf_file or pdf_file.filename == '':
        flash('Please select a PDF file.', 'error')
        return redirect(url_for('upload'))
    try:
        file_bytes = pdf_file.read()
    except Exception:
        flash('Could not read the uploaded file.', 'error')
        return redirect(url_for('upload'))
    text = extract_pdf_text(file_bytes, password if password else None)
    if not text:
        flash('Could not read the PDF. Please check the password and try again.', 'error')
        return redirect(url_for('upload'))
    holdings = parse_cdsl_pdf(text)
    if not holdings:
        flash('No stock holdings found in this PDF.', 'error')
        return redirect(url_for('upload'))
    Stock.query.filter_by(user_id=current_user.id).delete()
    live_fetched = 0
    for h in holdings:
        ticker, live_price = fetch_live_price_by_isin(h['isin'], h['name'])
        if live_price:
            value = round(h['quantity'] * live_price, 2)
            live_fetched += 1
        else:
            live_price = h['price']
            value = h['value']
        stock = Stock(user_id=current_user.id, isin=h['isin'], name=h['name'],
            quantity=h['quantity'], buy_price=h['price'], live_price=live_price,
            value=value, ticker=ticker, source='cdsl',
            price_updated_at=dt.utcnow() if live_price else None)
        db.session.add(stock)
    db.session.commit()
    flash(f'Imported {len(holdings)} stocks. Live prices fetched for {live_fetched}.', 'success')
    return redirect(url_for('upload'))

@app.route('/upload/delete-mf', methods=['POST'])
@login_required
def delete_all_mf():
    MutualFund.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash('All mutual fund data cleared.', 'success')
    return redirect(url_for('upload'))

@app.route('/upload/delete-stocks', methods=['POST'])
@login_required
def delete_all_stocks():
    Stock.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash('All stock data cleared.', 'success')
    return redirect(url_for('upload'))

def calculate_goal(target_amt, target_year, current_savings, monthly_sip, annual_return, inflation_rate=0):
    import math as _math
    from datetime import datetime as _dt
    current_year = _dt.now().year
    years  = max(target_year - current_year, 0)
    months = max(years * 12, 1)
    r = (annual_return / 100) / 12

    # Inflation-adjusted target
    if inflation_rate and inflation_rate > 0:
        inflation_adjusted_target = round(target_amt * ((1 + inflation_rate / 100) ** years), 2)
    else:
        inflation_adjusted_target = target_amt

    if r > 0:
        fv_sip = monthly_sip * (((1 + r) ** months - 1) / r) * (1 + r)
    else:
        fv_sip = monthly_sip * months
    fv_savings = current_savings * ((1 + r) ** months)
    projected  = round(fv_sip + fv_savings, 2)

    shortfall = round(inflation_adjusted_target - projected, 2)
    if r > 0 and shortfall > 0:
        required_sip = round((shortfall * r) / (((1 + r) ** months - 1) * (1 + r)), 2)
    else:
        required_sip = 0
    progress   = min(round((projected / inflation_adjusted_target) * 100, 1), 100) if inflation_adjusted_target > 0 else 0
    years_left = target_year - current_year
    return {
        'projected':                  projected,
        'shortfall':                  shortfall,
        'surplus':                    max(-shortfall, 0),
        'on_track':                   shortfall <= 0,
        'required_sip':               required_sip,
        'progress':                   progress,
        'years_left':                 years_left,
        'months':                     months,
        'inflation_adjusted_target':  inflation_adjusted_target,
        'inflation_applied':          inflation_rate > 0,
    }

@app.route('/goals')
@login_required
def goals():
    user_goals = Goal.query.filter_by(user_id=current_user.id).order_by(Goal.target_year).all()
    goals_data = []
    for g in user_goals:
        calc = calculate_goal(g.target_amt, g.target_year, g.current_savings, g.monthly_sip, g.annual_return, getattr(g, 'inflation_rate', 0) or 0)
        goals_data.append({'goal': g, 'calc': calc})
    return render_template('goals.html', goals_data=goals_data)

@app.route('/goals/add', methods=['POST'])
@login_required
def add_goal():
    name        = request.form.get('name', '').strip()
    emoji       = request.form.get('emoji', '').strip()
    target_amt  = safe_float(request.form.get('target_amt'))
    target_year = int(request.form.get('target_year', 2030))
    current_savings = safe_float(request.form.get('current_savings'))
    monthly_sip     = safe_float(request.form.get('monthly_sip'))
    annual_return   = safe_float(request.form.get('annual_return', '12'))
    if not name or target_amt <= 0:
        flash('Please enter a goal name and target amount.', 'error')
        return redirect(url_for('goals'))
    if target_year <= 2024:
        flash('Target year must be in the future.', 'error')
        return redirect(url_for('goals'))
    inflation_rate = safe_float(request.form.get('inflation_rate', '0'))
    goal = Goal(
        user_id=current_user.id, name=name, emoji=emoji,
        target_amt=target_amt, target_year=target_year,
        current_savings=current_savings, monthly_sip=monthly_sip,
        annual_return=annual_return if annual_return > 0 else 12.0,
        inflation_rate=inflation_rate if inflation_rate >= 0 else 0
    )
    db.session.add(goal)
    db.session.commit()
    flash(f'Goal added successfully!', 'success')
    return redirect(url_for('goals'))

@app.route('/goals/delete/<int:goal_id>', methods=['POST'])
@login_required
def delete_goal(goal_id):
    goal = Goal.query.get_or_404(goal_id)
    if goal.user_id != current_user.id:
        flash('Permission denied.', 'error')
        return redirect(url_for('goals'))
    db.session.delete(goal)
    db.session.commit()
    flash('Goal deleted.', 'success')
    return redirect(url_for('goals'))


# ── Step 7: Export routes (PDF + Excel) ──────────────────────────────────────
from flask import send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable, KeepTogether)
from reportlab.platypus import PageBreak
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             GradientFill)
from openpyxl.utils import get_column_letter
import io as _io
from datetime import datetime as _dt

# ── shared colours ────────────────────────────────────────────────────────────
_TEAL   = colors.HexColor('#00d4aa')
_DARK   = colors.HexColor('#0d0f14')
_CARD   = colors.HexColor('#13161f')
_BORDER = colors.HexColor('#1e2130')
_MUTED  = colors.HexColor('#64748b')
_WHITE  = colors.white
_RED    = colors.HexColor('#ef4444')
_GREEN  = colors.HexColor('#22c55e')
_AMBER  = colors.HexColor('#f59e0b')

# ── PDF helpers ───────────────────────────────────────────────────────────────
def _fmt(n):
    return f"₹{n:,.0f}"

def _pct(n):
    return f"{n:.1f}%"

def _pdf_styles():
    base = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, **kw)
    return {
        'title':     S('title',   fontSize=22, textColor=_TEAL,  leading=28, fontName='Helvetica-Bold'),
        'subtitle':  S('sub',     fontSize=9,  textColor=_MUTED, leading=14, fontName='Helvetica'),
        'h2':        S('h2',      fontSize=13, textColor=_WHITE, leading=18, fontName='Helvetica-Bold', spaceAfter=4),
        'h3':        S('h3',      fontSize=10, textColor=_TEAL,  leading=14, fontName='Helvetica-Bold', spaceAfter=2),
        'body':      S('body',    fontSize=9,  textColor=_WHITE, leading=13, fontName='Helvetica'),
        'muted':     S('muted',   fontSize=8,  textColor=_MUTED, leading=12, fontName='Helvetica'),
        'right':     S('right',   fontSize=9,  textColor=_WHITE, leading=13, fontName='Helvetica', alignment=TA_RIGHT),
        'teal_right':S('tr',      fontSize=9,  textColor=_TEAL,  leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'green':     S('green',   fontSize=9,  textColor=_GREEN, leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'red':       S('red',     fontSize=9,  textColor=_RED,   leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'amber':     S('amber',   fontSize=9,  textColor=_AMBER, leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
    }

def _tbl_style(header_bg=None, row_colors=True):
    hbg = header_bg or _CARD
    cmds = [
        ('BACKGROUND',   (0, 0), (-1, 0),  hbg),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  _TEAL),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0),  8),
        ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('TEXTCOLOR',    (0, 1), (-1, -1), _WHITE),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
            [colors.HexColor('#13161f'), colors.HexColor('#1a1e2d')] if row_colors else [_CARD]),
        ('GRID',         (0, 0), (-1, -1), 0.3, _BORDER),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',   (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
    ]
    return TableStyle(cmds)

def _section_header(text, styles):
    return [
        Spacer(1, 6*mm),
        HRFlowable(width='100%', thickness=0.5, color=_TEAL, spaceAfter=3),
        Paragraph(text, styles['h2']),
        Spacer(1, 2*mm),
    ]

def _sip_projection(target_amt, target_year, current_savings, monthly_sip, annual_return):
    """Return list of (month_label, balance) for 12 months + yearly to target."""
    rows = []
    r = annual_return / 100 / 12
    balance = float(current_savings or 0)
    now = _dt.now()
    cur_year = now.year
    cur_month = now.month

    # Monthly for first 12 months
    for i in range(1, 13):
        if r > 0:
            balance = balance * (1 + r) + float(monthly_sip or 0)
        else:
            balance += float(monthly_sip or 0)
        m = (cur_month + i - 1) % 12 + 1
        y = cur_year + (cur_month + i - 1) // 12
        rows.append((f"{_dt(y, m, 1).strftime('%b %Y')}", balance, "monthly"))

    # Yearly milestones after month 12
    balance_after_12 = rows[-1][1]
    months_done = 12
    total_months = (target_year - cur_year) * 12 - cur_month + 1
    for yr in range(cur_year + 1, target_year + 1):
        months_to_yr = (yr - cur_year) * 12 - cur_month + 1
        if months_to_yr <= 12:
            continue
        b = balance_after_12
        extra = months_to_yr - 12
        for _ in range(extra):
            if r > 0:
                b = b * (1 + r) + float(monthly_sip or 0)
            else:
                b += float(monthly_sip or 0)
        rows.append((f"Dec {yr}", b, "yearly"))

    return rows

# ── PDF EXPORT ────────────────────────────────────────────────────────────────
@app.route('/export/pdf')
@login_required
def export_pdf():
    # ── gather data ──
    # Phase H: asset figures now come from WealthAsset (via
    # WealthStatisticsService), the authoritative Wealth source,
    # instead of the retired legacy Asset model.
    wstats  = WealthStatisticsService(current_user.id)
    assets_by_cat = wstats.assets_by_category()
    mfs     = MutualFund.query.filter_by(user_id=current_user.id).all()
    stocks  = Stock.query.filter_by(user_id=current_user.id).all()
    goals   = Goal.query.filter_by(user_id=current_user.id).order_by(Goal.target_year).all()

    equity_val     = sum(m.value for m in mfs) + sum(s.value for s in stocks)
    debt_val       = (sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.BANK_DEPOSITS, []))
                       + sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.INVESTMENTS, [])))
    gold_val       = sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.PRECIOUS_METALS, []))
    realestate_val = sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.REAL_ESTATE, []))
    other_val      = (sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.VEHICLES, []))
                       + sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.BUSINESS, []))
                       + sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.OTHER, [])))
    cash_val       = 0
    total          = equity_val + debt_val + gold_val + realestate_val + cash_val + other_val

    # ── build PDF ──
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
        title="MyWealthLens — Wealth Report",
    )
    W = doc.width
    S = _pdf_styles()
    story = []

    # ── COVER HEADER ──
    now_str = _dt.now().strftime("%d %B %Y, %I:%M %p")
    story += [
        Paragraph("MyWealthLens", S['title']),
        Paragraph("Personal Wealth Report", S['h2']),
        Paragraph(f"Generated on {now_str}", S['muted']),
        Spacer(1, 4*mm),
    ]

    # ── USER PROFILE CARD ──
    story += _section_header("👤  User Profile", S)
    profile_data = [
        ['Name', current_user.name, 'Email', current_user.email],
    ]

    pt = Table(profile_data, colWidths=[W*0.18, W*0.32, W*0.18, W*0.32])
    pt.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,-1), _CARD),
        ('TEXTCOLOR',   (0,0), (-1,-1), _WHITE),
        ('TEXTCOLOR',   (0,0), (0,-1),  _MUTED),
        ('TEXTCOLOR',   (2,0), (2,-1),  _MUTED),
        ('FONTNAME',    (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,0), (-1,-1), 9),
        ('FONTNAME',    (1,0), (1,-1),  'Helvetica-Bold'),
        ('FONTNAME',    (3,0), (3,-1),  'Helvetica-Bold'),
        ('GRID',        (0,0), (-1,-1), 0.3, _BORDER),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING',  (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
    ]))
    story += [pt, Spacer(1, 4*mm)]

    # ── NET WORTH SUMMARY ──
    story += _section_header("💰  Net Worth Summary", S)
    story.append(Paragraph(f"Total Net Worth: {_fmt(total)}", ParagraphStyle(
        'big', fontSize=16, textColor=_TEAL, fontName='Helvetica-Bold', spaceAfter=6)))

    def pct(v): return round(v/total*100, 1) if total else 0

    summary_data = [
        ['Asset Class', 'Value (₹)', 'Allocation %'],
        ['📊 Equity (MF + Stocks)', _fmt(equity_val), _pct(pct(equity_val))],
        ['🏛️ Debt (PPF/VPF/SSY/FD)', _fmt(debt_val),  _pct(pct(debt_val))],
        ['🥇 Gold / Silver',          _fmt(gold_val),  _pct(pct(gold_val))],
        ['🏠 Real Estate',             _fmt(realestate_val), _pct(pct(realestate_val))],
        ['💵 Cash & Others',           _fmt(cash_val + other_val), _pct(pct(cash_val + other_val))],
        ['TOTAL', _fmt(total), '100%'],
    ]
    st = Table(summary_data, colWidths=[W*0.52, W*0.26, W*0.22])
    st.setStyle(_tbl_style())
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,-1), (-1,-1), _TEAL),
        ('TEXTCOLOR',  (0,-1), (-1,-1), _DARK),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN',      (1,0),  (-1,-1), 'RIGHT'),
    ]))
    story += [st, Spacer(1, 2*mm)]

    # ── ASSET DETAIL TABLES ──
    story += _section_header("📋  Asset Details", S)

    # Mutual Funds
    if mfs:
        story.append(Paragraph("Mutual Funds", S['h3']))
        mf_data = [['Scheme Name', 'Folio', 'Units', 'NAV (₹)', 'Value (₹)']]
        for m in mfs:
            mf_data.append([
                m.scheme_name or m.name or '—',
                getattr(m, 'folio', '—') or '—',
                f"{getattr(m, 'units', 0) or 0:,.3f}",
                f"{getattr(m, 'nav', 0) or 0:,.2f}",
                _fmt(m.value),
            ])
        mt = Table(mf_data, colWidths=[W*0.42, W*0.16, W*0.12, W*0.14, W*0.16])
        mt.setStyle(_tbl_style())
        story += [mt, Spacer(1, 3*mm)]

    # Stocks
    if stocks:
        story.append(Paragraph("Stocks / Demat Holdings", S['h3']))
        sk_data = [['Company / ISIN', 'Quantity', 'Price (₹)', 'Value (₹)']]
        for s in stocks:
            sk_data.append([
                s.name or getattr(s, 'isin', '—') or '—',
                f"{getattr(s, 'quantity', 0) or 0:,.0f}",
                f"{getattr(s, 'price', 0) or 0:,.2f}",
                _fmt(s.value),
            ])
        skt = Table(sk_data, colWidths=[W*0.46, W*0.16, W*0.18, W*0.20])
        skt.setStyle(_tbl_style())
        story += [skt, Spacer(1, 3*mm)]

    # Physical / other assets — itemized by Wealth Asset category
    cat_icons = {
        WealthAssetCategory.REAL_ESTATE:     '🏠',
        WealthAssetCategory.PRECIOUS_METALS: '🥇',
        WealthAssetCategory.VEHICLES:        '🚗',
        WealthAssetCategory.BANK_DEPOSITS:   '🏦',
        WealthAssetCategory.INVESTMENTS:     '📈',
        WealthAssetCategory.BUSINESS:        '🏢',
        WealthAssetCategory.OTHER:           '📦',
    }

    for cat in WealthAssetCategory.ALL:
        items = assets_by_cat.get(cat, [])
        if not items:
            continue
        story.append(Paragraph(f"{cat_icons.get(cat, '📦')} {cat}", S['h3']))
        ph_data = [['Name', 'Value (₹)']]
        for a in items:
            ph_data.append([a.name or cat, _fmt(a.current_value)])
        ph_data.append(['Subtotal', _fmt(sum(a.current_value for a in items))])
        pht = Table(ph_data, colWidths=[W*0.70, W*0.30])
        pht.setStyle(_tbl_style())
        pht.setStyle(TableStyle([
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#1a1e2d')),
            ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN',      (1,0),  (1,-1),  'RIGHT'),
        ]))
        story += [pht, Spacer(1, 3*mm)]

    # ── GOALS + SIP PROJECTIONS ──
    if goals:
        story += _section_header("🎯  Financial Goals & SIP Projections", S)
        for g in goals:
            calc = calculate_goal(g.target_amt, g.target_year,
                                  g.current_savings, g.monthly_sip, g.annual_return)
            status_color = _GREEN if calc['on_track'] else _RED
            story.append(KeepTogether([
                Paragraph(f"{g.emoji or '⭐'} {g.name}", S['h3']),
                Spacer(1, 1*mm),
            ]))
            g_meta = [
                ['Target Amount', _fmt(g.target_amt), 'Target Year', str(g.target_year)],
                ['Current Savings', _fmt(g.current_savings or 0),
                 'Monthly SIP', _fmt(g.monthly_sip or 0)],
                ['Annual Return', f"{g.annual_return}% p.a.",
                 'Projected Corpus', _fmt(calc['projected'])],
                ['Years Left', str(calc['years_left']),
                 'Status',
                 'On Track ✓' if calc['on_track'] else f"Shortfall ₹{calc.get('shortfall',0):,.0f}"],
            ]
            gmt = Table(g_meta, colWidths=[W*0.22, W*0.28, W*0.22, W*0.28])
            gmt.setStyle(TableStyle([
                ('BACKGROUND',   (0,0), (-1,-1), _CARD),
                ('TEXTCOLOR',    (0,0), (-1,-1), _WHITE),
                ('TEXTCOLOR',    (0,0), (0,-1),  _MUTED),
                ('TEXTCOLOR',    (2,0), (2,-1),  _MUTED),
                ('FONTNAME',     (1,0), (1,-1),  'Helvetica-Bold'),
                ('FONTNAME',     (3,0), (3,-1),  'Helvetica-Bold'),
                ('FONTSIZE',     (0,0), (-1,-1),  8),
                ('GRID',         (0,0), (-1,-1),  0.3, _BORDER),
                ('LEFTPADDING',  (0,0), (-1,-1),  6),
                ('TOPPADDING',   (0,0), (-1,-1),  5),
                ('BOTTOMPADDING',(0,0), (-1,-1),  5),
            ]))
            story += [gmt, Spacer(1, 2*mm)]

            # SIP projection table
            proj = _sip_projection(g.target_amt, g.target_year,
                                   g.current_savings, g.monthly_sip, g.annual_return)
            if proj:
                story.append(Paragraph("SIP Growth Projection", S['h3']))
                sip_hdr = [['Period', 'Projected Balance (₹)', 'vs Target (₹)', 'Progress %']]
                sip_rows = []
                for (label, bal, kind) in proj:
                    delta = bal - g.target_amt
                    progress = min(round(bal / g.target_amt * 100, 1), 100) if g.target_amt else 0
                    sip_rows.append([
                        label,
                        f"{bal:,.0f}",
                        f"{'+' if delta >= 0 else ''}{delta:,.0f}",
                        f"{progress}%",
                    ])
                sip_data = sip_hdr + sip_rows
                sipt = Table(sip_data, colWidths=[W*0.22, W*0.28, W*0.28, W*0.22])
                sipt.setStyle(_tbl_style())
                sipt.setStyle(TableStyle([('ALIGN',(1,0),(-1,-1),'RIGHT')]))
                story += [sipt, Spacer(1, 5*mm)]

    # ── FOOTER ──
    story += [
        HRFlowable(width='100%', thickness=0.5, color=_BORDER),
        Spacer(1, 2*mm),
        Paragraph("Generated by MyWealthLens · For personal use only · Not investment advice",
                  ParagraphStyle('footer', fontSize=7, textColor=_MUTED,
                                 fontName='Helvetica', alignment=TA_CENTER)),
    ]

    # ── page background ──
    def _dark_bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(_DARK)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=_dark_bg, onLaterPages=_dark_bg)
    buf.seek(0)
    fname = f"MyWealthLens_{current_user.name.replace(' ','_')}_{_dt.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=fname)


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
@app.route('/export/excel')
@login_required
def export_excel():
    # Phase H: sourced from WealthAsset (authoritative), not the
    # retired legacy Asset model.
    wstats  = WealthStatisticsService(current_user.id)
    assets_by_cat = wstats.assets_by_category()
    all_assets_flat = [a for items in assets_by_cat.values() for a in items]
    mfs     = MutualFund.query.filter_by(user_id=current_user.id).all()
    stocks  = Stock.query.filter_by(user_id=current_user.id).all()
    goals   = Goal.query.filter_by(user_id=current_user.id).order_by(Goal.target_year).all()

    equity_val     = sum(m.value for m in mfs) + sum(s.value for s in stocks)
    debt_val       = (sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.BANK_DEPOSITS, []))
                       + sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.INVESTMENTS, [])))
    gold_val       = sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.PRECIOUS_METALS, []))
    realestate_val = sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.REAL_ESTATE, []))
    other_val      = (sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.VEHICLES, []))
                       + sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.BUSINESS, []))
                       + sum(a.current_value for a in assets_by_cat.get(WealthAssetCategory.OTHER, [])))
    cash_val       = 0
    total          = equity_val + debt_val + gold_val + realestate_val + cash_val + other_val

    wb = openpyxl.Workbook()

    # ── colour helpers ──
    TEAL_HEX  = '00D4AA'
    DARK_HEX  = '0D0F14'
    CARD_HEX  = '13161F'
    CARD2_HEX = '1A1E2D'
    BORD_HEX  = '1E2130'
    WHITE_HEX = 'E2E8F0'
    MUTED_HEX = '64748B'
    GREEN_HEX = '22C55E'
    RED_HEX   = 'EF4444'
    AMBER_HEX = 'F59E0B'

    def _fill(hex_): return PatternFill('solid', fgColor=hex_)
    def _font(hex_=WHITE_HEX, bold=False, sz=10):
        return Font(color=hex_, bold=bold, size=sz, name='Calibri')
    def _border():
        s = Side(style='thin', color=BORD_HEX)
        return Border(left=s, right=s, top=s, bottom=s)
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _hdr_row(ws, row_num, values, col_start=1):
        for i, v in enumerate(values):
            c = ws.cell(row=row_num, column=col_start+i, value=v)
            c.fill = _fill(CARD_HEX)
            c.font = _font(TEAL_HEX, bold=True, sz=9)
            c.border = _border()
            c.alignment = _align('center')

    def _data_row(ws, row_num, values, col_start=1, alt=False):
        bg = CARD2_HEX if alt else CARD_HEX
        for i, v in enumerate(values):
            c = ws.cell(row=row_num, column=col_start+i, value=v)
            c.fill = _fill(bg)
            c.font = _font(sz=9)
            c.border = _border()
            c.alignment = _align()

    def _title_cell(ws, row_num, col, text, sz=14):
        c = ws.cell(row=row_num, column=col, value=text)
        c.font = Font(color=TEAL_HEX, bold=True, size=sz, name='Calibri')
        c.fill = _fill(DARK_HEX)
        c.alignment = _align()

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _freeze(ws, cell='A2'):
        ws.freeze_panes = cell

    def _tab_color(ws, hex_):
        ws.sheet_properties.tabColor = hex_

    # ════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    _tab_color(ws1, TEAL_HEX)
    ws1.sheet_view.showGridLines = False
    for row in ws1.iter_rows(min_row=1, max_row=60, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws1, 1, 1, "MyWealthLens — Wealth Report", sz=16)
    ws1.cell(row=2, column=1, value=f"Generated: {_dt.now().strftime('%d %B %Y, %I:%M %p')}").font = _font(MUTED_HEX, sz=9)

    r = 4
    _title_cell(ws1, r, 1, "USER PROFILE", sz=11)
    r += 1
    _hdr_row(ws1, r, ['Field', 'Value'])
    r += 1
    profile_rows = [
        ('Name', current_user.name),
        ('Email', current_user.email),
    ]
    for i, (k, v) in enumerate(profile_rows):
        _data_row(ws1, r, [k, v], alt=i%2==1)
        r += 1

    r += 1
    _title_cell(ws1, r, 1, "NET WORTH SUMMARY", sz=11)
    r += 1
    _hdr_row(ws1, r, ['Asset Class', 'Value (₹)', 'Allocation %'])
    r += 1
    summary_rows = [
        ('Equity (MF + Stocks)', equity_val),
        ('Debt (PPF/VPF/SSY/FD)', debt_val),
        ('Gold / Silver', gold_val),
        ('Real Estate', realestate_val),
        ('Cash & Others', cash_val + other_val),
    ]
    for i, (lbl, val) in enumerate(summary_rows):
        pct_val = round(val/total*100, 1) if total else 0
        _data_row(ws1, r, [lbl, val, f"{pct_val}%"], alt=i%2==1)
        ws1.cell(row=r, column=2).number_format = '₹#,##0'
        r += 1
    # Total row
    for col, val in enumerate([('TOTAL', total, '100%')], 1):
        pass
    _hdr_row(ws1, r, ['TOTAL', total, '100%'])
    ws1.cell(row=r, column=2).number_format = '₹#,##0'

    _set_col_widths(ws1, [28, 20, 15])
    _freeze(ws1, 'A5')

    # ════════════════════════════════════════════════
    # SHEET 2 — All Assets
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Assets")
    _tab_color(ws2, '6366F1')
    ws2.sheet_view.showGridLines = False
    for row in ws2.iter_rows(min_row=1, max_row=500, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws2, 1, 1, "All Assets", sz=14)
    r2 = 3
    _hdr_row(ws2, r2, ['Category', 'Name / Scheme', 'Sub-type', 'Units / Qty', 'Price / NAV (₹)', 'Value (₹)'])
    r2 += 1

    all_rows = []
    for m in mfs:
        all_rows.append(['Mutual Fund', m.scheme_name or m.name or '—',
                         getattr(m, 'amc', '') or '',
                         getattr(m, 'units', '') or '', getattr(m, 'nav', '') or '', m.value])
    for s in stocks:
        all_rows.append(['Stock', s.name or '—', getattr(s, 'isin', '') or '',
                         getattr(s, 'quantity', '') or '', getattr(s, 'price', '') or '', s.value])
    for a in all_assets_flat:
        all_rows.append([a.category, a.name or '—', a.asset_type or '', '', '', a.current_value])

    for i, row_data in enumerate(all_rows):
        _data_row(ws2, r2, row_data, alt=i%2==1)
        ws2.cell(row=r2, column=6).number_format = '₹#,##0'
        r2 += 1

    _set_col_widths(ws2, [16, 40, 20, 12, 15, 16])
    _freeze(ws2, 'A4')

    # ════════════════════════════════════════════════
    # SHEET 3 — Goals + SIP Projections
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet("Goals & Projections")
    _tab_color(ws4, 'F59E0B')
    ws4.sheet_view.showGridLines = False
    for row in ws4.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws4, 1, 1, "Goals & SIP Projections", sz=14)
    r4 = 3

    if goals:
        for g in goals:
            calc = calculate_goal(g.target_amt, g.target_year,
                                  g.current_savings, g.monthly_sip, g.annual_return)
            _title_cell(ws4, r4, 1, f"{g.emoji or '⭐'} {g.name}", sz=12)
            r4 += 1
            _hdr_row(ws4, r4, ['Target (₹)', 'Target Year', 'Savings (₹)',
                                'SIP/mo (₹)', 'Return %', 'Projected (₹)', 'Status'])
            r4 += 1
            status = 'On Track ✓' if calc['on_track'] else f"Shortfall ₹{calc.get('shortfall',0):,.0f}"
            _data_row(ws4, r4, [g.target_amt, g.target_year, g.current_savings or 0,
                                 g.monthly_sip or 0, f"{g.annual_return}%",
                                 round(calc['projected']), status])
            for col in [1, 3, 4, 6]:
                ws4.cell(row=r4, column=col).number_format = '₹#,##0'
            ws4.cell(row=r4, column=7).font = \
                _font(GREEN_HEX if calc['on_track'] else RED_HEX, bold=True, sz=9)
            r4 += 2

            # SIP projection
            ws4.cell(row=r4, column=1, value="SIP Growth Projection").font = _font(TEAL_HEX, bold=True, sz=10)
            r4 += 1
            _hdr_row(ws4, r4, ['Period', 'Projected Balance (₹)', 'vs Target (₹)', 'Progress %'])
            r4 += 1

            proj = _sip_projection(g.target_amt, g.target_year,
                                   g.current_savings, g.monthly_sip, g.annual_return)
            for i, (label, bal, _) in enumerate(proj):
                delta = bal - g.target_amt
                progress = min(round(bal / g.target_amt * 100, 1), 100) if g.target_amt else 0
                _data_row(ws4, r4, [label, round(bal), round(delta), f"{progress}%"], alt=i%2==1)
                ws4.cell(row=r4, column=2).number_format = '₹#,##0'
                ws4.cell(row=r4, column=3).number_format = '₹#,##0'
                delta_color = GREEN_HEX if delta >= 0 else RED_HEX
                ws4.cell(row=r4, column=3).font = _font(delta_color, sz=9)
                r4 += 1
            r4 += 2
    else:
        ws4.cell(row=r4, column=1, value="No goals configured.").font = _font(MUTED_HEX)

    _set_col_widths(ws4, [16, 20, 18, 16])
    _freeze(ws4, 'A3')

    # ── save + send ──
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"MyWealthLens_{current_user.name.replace(' ','_')}_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)



@app.errorhandler(429)
def rate_limit_exceeded(e):
    if request.method == 'GET':
        return render_template('login.html',
            error='Too many login attempts. Please wait 15 minutes.'), 429
    flash('Too many attempts. Please wait 15 minutes before trying again.', 'error')
    return redirect(url_for('login')), 429

@app.route('/account/change-password', methods=['POST'])
@login_required
def change_password():
    current_pw = request.form.get('current_password', '')
    new_pw     = request.form.get('new_password', '')
    confirm_pw = request.form.get('confirm_password', '')
    if not bcrypt.checkpw(current_pw.encode('utf-8'), current_user.password.encode('utf-8')):
        flash('Current password is incorrect.', 'error')
        return redirect(url_for('preferences'))
    if len(new_pw) < 8:
        flash('New password must be at least 8 characters.', 'error')
        return redirect(url_for('preferences'))
    if new_pw != confirm_pw:
        flash('New passwords do not match.', 'error')
        return redirect(url_for('preferences'))
    hashed = bcrypt.hashpw(new_pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    current_user.password = hashed
    db.session.commit()
    flash('Password changed successfully!', 'success')
    return redirect(url_for('preferences'))

app.register_blueprint(insurance_bp)
app.register_blueprint(retirement_bp)
app.register_blueprint(wealth_bp)

# Phase I — Automatic Wealth Snapshots. Registers `flask wealth
# snapshot`, invoked by Windows Task Scheduler (see the Phase I
# final report for setup). Kept as a separate module rather than
# defined inline here, matching this project's existing pattern of
# routes.py/services.py living inside each module's own folder.
from wealth.cli import register_cli
register_cli(app)

if __name__ == '__main__': 
     app.run(debug=True, port=5000, host='0.0.0.0')