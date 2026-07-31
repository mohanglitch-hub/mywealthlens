# ── Step 7: Export routes (PDF + Excel) ──────────────────────────────────────
from flask import send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable, KeepTogether)
from reportlab.platypus import PageBreak
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             GradientFill)
from openpyxl.utils import get_column_letter
import io as _io
from datetime import datetime as _dt

# ── shared colours ────────────────────────────────────────────────────────────
_TEAL   = colors.HexColor('#00d4aa')
_DARK   = colors.HexColor('#0d0f14')
_CARD   = colors.HexColor('#13161f')
_BORDER = colors.HexColor('#1e2130')
_MUTED  = colors.HexColor('#64748b')
_WHITE  = colors.white
_RED    = colors.HexColor('#ef4444')
_GREEN  = colors.HexColor('#22c55e')
_AMBER  = colors.HexColor('#f59e0b')

# ── PDF helpers ───────────────────────────────────────────────────────────────
def _fmt(n):
    return f"₹{n:,.0f}"

def _pct(n):
    return f"{n:.1f}%"

def _pdf_styles():
    base = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, **kw)
    return {
        'title':     S('title',   fontSize=22, textColor=_TEAL,  leading=28, fontName='Helvetica-Bold'),
        'subtitle':  S('sub',     fontSize=9,  textColor=_MUTED, leading=14, fontName='Helvetica'),
        'h2':        S('h2',      fontSize=13, textColor=_WHITE, leading=18, fontName='Helvetica-Bold', spaceAfter=4),
        'h3':        S('h3',      fontSize=10, textColor=_TEAL,  leading=14, fontName='Helvetica-Bold', spaceAfter=2),
        'body':      S('body',    fontSize=9,  textColor=_WHITE, leading=13, fontName='Helvetica'),
        'muted':     S('muted',   fontSize=8,  textColor=_MUTED, leading=12, fontName='Helvetica'),
        'right':     S('right',   fontSize=9,  textColor=_WHITE, leading=13, fontName='Helvetica', alignment=TA_RIGHT),
        'teal_right':S('tr',      fontSize=9,  textColor=_TEAL,  leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'green':     S('green',   fontSize=9,  textColor=_GREEN, leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'red':       S('red',     fontSize=9,  textColor=_RED,   leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'amber':     S('amber',   fontSize=9,  textColor=_AMBER, leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT),
    }

def _tbl_style(header_bg=None, row_colors=True):
    hbg = header_bg or _CARD
    cmds = [
        ('BACKGROUND',   (0, 0), (-1, 0),  hbg),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  _TEAL),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0),  8),
        ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('TEXTCOLOR',    (0, 1), (-1, -1), _WHITE),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
            [colors.HexColor('#13161f'), colors.HexColor('#1a1e2d')] if row_colors else [_CARD]),
        ('GRID',         (0, 0), (-1, -1), 0.3, _BORDER),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',   (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
    ]
    return TableStyle(cmds)

def _section_header(text, styles):
    return [
        Spacer(1, 6*mm),
        HRFlowable(width='100%', thickness=0.5, color=_TEAL, spaceAfter=3),
        Paragraph(text, styles['h2']),
        Spacer(1, 2*mm),
    ]

def _sip_projection(target_amt, target_year, current_savings, monthly_sip, annual_return):
    """Return list of (month_label, balance) for 12 months + yearly to target."""
    rows = []
    r = annual_return / 100 / 12
    balance = float(current_savings or 0)
    now = _dt.now()
    cur_year = now.year
    cur_month = now.month

    # Monthly for first 12 months
    for i in range(1, 13):
        if r > 0:
            balance = balance * (1 + r) + float(monthly_sip or 0)
        else:
            balance += float(monthly_sip or 0)
        m = (cur_month + i - 1) % 12 + 1
        y = cur_year + (cur_month + i - 1) // 12
        rows.append((f"{_dt(y, m, 1).strftime('%b %Y')}", balance, "monthly"))

    # Yearly milestones after month 12
    balance_after_12 = rows[-1][1]
    months_done = 12
    total_months = (target_year - cur_year) * 12 - cur_month + 1
    for yr in range(cur_year + 1, target_year + 1):
        months_to_yr = (yr - cur_year) * 12 - cur_month + 1
        if months_to_yr <= 12:
            continue
        b = balance_after_12
        extra = months_to_yr - 12
        for _ in range(extra):
            if r > 0:
                b = b * (1 + r) + float(monthly_sip or 0)
            else:
                b += float(monthly_sip or 0)
        rows.append((f"Dec {yr}", b, "yearly"))

    return rows

# ── PDF EXPORT ────────────────────────────────────────────────────────────────
@app.route('/export/pdf')
@login_required
def export_pdf():
    # ── gather data ──
    assets  = Asset.query.filter_by(user_id=current_user.id).all()
    mfs     = MutualFund.query.filter_by(user_id=current_user.id).all()
    stocks  = Stock.query.filter_by(user_id=current_user.id).all()
    goals   = Goal.query.filter_by(user_id=current_user.id).order_by(Goal.target_year).all()
    profile = UserProfile.query.filter_by(user_id=current_user.id).first()

    equity_val     = sum(m.value for m in mfs) + sum(s.value for s in stocks)
    debt_val       = sum(a.value for a in assets if a.category in ['ppf','vpf','ssy','fd'])
    gold_val       = sum(a.value for a in assets if a.category in ['gold','silver'])
    realestate_val = sum(a.value for a in assets if a.category.startswith('real_estate'))
    cash_val       = sum(a.value for a in assets if a.category == 'cash')
    other_val      = sum(a.value for a in assets if a.category == 'other')
    total          = equity_val + debt_val + gold_val + realestate_val + cash_val + other_val

    stage = stage_data = allocation = recommendations = health_score = None
    if profile:
        stage_key = detect_life_stage(profile.age, profile.marital_status, profile.dependents)
        stage_data = LIFE_STAGES[stage_key]
        _, allocation, recommendations, _, health_score, _ = \
            compute_life_stage_data(profile, assets, mfs, stocks)

    # ── build PDF ──
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
        title="MyWealthLens — Wealth Report",
    )
    W = doc.width
    S = _pdf_styles()
    story = []

    # ── COVER HEADER ──
    now_str = _dt.now().strftime("%d %B %Y, %I:%M %p")
    story += [
        Paragraph("MyWealthLens", S['title']),
        Paragraph("Personal Wealth Report", S['h2']),
        Paragraph(f"Generated on {now_str}", S['muted']),
        Spacer(1, 4*mm),
    ]

    # ── USER PROFILE CARD ──
    story += _section_header("👤  User Profile", S)
    profile_data = [
        ['Name', current_user.name, 'Email', current_user.email],
    ]
    if profile:
        stage_label = stage_data['name'] if stage_data else '—'
        profile_data += [
            ['Age', str(profile.age),
             'Life Stage', f"{stage_data['icon']} {stage_label}" if stage_data else '—'],
            ['Marital Status', profile.marital_status.title(),
             'Dependents', str(profile.dependents)],
        ]
    else:
        profile_data.append(['Life Stage Profile', 'Not configured', '', ''])

    pt = Table(profile_data, colWidths=[W*0.18, W*0.32, W*0.18, W*0.32])
    pt.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,-1), _CARD),
        ('TEXTCOLOR',   (0,0), (-1,-1), _WHITE),
        ('TEXTCOLOR',   (0,0), (0,-1),  _MUTED),
        ('TEXTCOLOR',   (2,0), (2,-1),  _MUTED),
        ('FONTNAME',    (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,0), (-1,-1), 9),
        ('FONTNAME',    (1,0), (1,-1),  'Helvetica-Bold'),
        ('FONTNAME',    (3,0), (3,-1),  'Helvetica-Bold'),
        ('GRID',        (0,0), (-1,-1), 0.3, _BORDER),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING',  (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
    ]))
    story += [pt, Spacer(1, 4*mm)]

    # ── NET WORTH SUMMARY ──
    story += _section_header("💰  Net Worth Summary", S)
    story.append(Paragraph(f"Total Net Worth: {_fmt(total)}", ParagraphStyle(
        'big', fontSize=16, textColor=_TEAL, fontName='Helvetica-Bold', spaceAfter=6)))

    def pct(v): return round(v/total*100, 1) if total else 0

    summary_data = [
        ['Asset Class', 'Value (₹)', 'Allocation %'],
        ['📊 Equity (MF + Stocks)', _fmt(equity_val), _pct(pct(equity_val))],
        ['🏛️ Debt (PPF/VPF/SSY/FD)', _fmt(debt_val),  _pct(pct(debt_val))],
        ['🥇 Gold / Silver',          _fmt(gold_val),  _pct(pct(gold_val))],
        ['🏠 Real Estate',             _fmt(realestate_val), _pct(pct(realestate_val))],
        ['💵 Cash & Others',           _fmt(cash_val + other_val), _pct(pct(cash_val + other_val))],
        ['TOTAL', _fmt(total), '100%'],
    ]
    st = Table(summary_data, colWidths=[W*0.52, W*0.26, W*0.22])
    st.setStyle(_tbl_style())
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,-1), (-1,-1), _TEAL),
        ('TEXTCOLOR',  (0,-1), (-1,-1), _DARK),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN',      (1,0),  (-1,-1), 'RIGHT'),
    ]))
    story += [st, Spacer(1, 2*mm)]

    # ── ASSET DETAIL TABLES ──
    story += _section_header("📋  Asset Details", S)

    # Mutual Funds
    if mfs:
        story.append(Paragraph("Mutual Funds", S['h3']))
        mf_data = [['Scheme Name', 'Folio', 'Units', 'NAV (₹)', 'Value (₹)']]
        for m in mfs:
            mf_data.append([
                m.scheme_name or m.name or '—',
                getattr(m, 'folio', '—') or '—',
                f"{getattr(m, 'units', 0) or 0:,.3f}",
                f"{getattr(m, 'nav', 0) or 0:,.2f}",
                _fmt(m.value),
            ])
        mt = Table(mf_data, colWidths=[W*0.42, W*0.16, W*0.12, W*0.14, W*0.16])
        mt.setStyle(_tbl_style())
        story += [mt, Spacer(1, 3*mm)]

    # Stocks
    if stocks:
        story.append(Paragraph("Stocks / Demat Holdings", S['h3']))
        sk_data = [['Company / ISIN', 'Quantity', 'Price (₹)', 'Value (₹)']]
        for s in stocks:
            sk_data.append([
                s.name or getattr(s, 'isin', '—') or '—',
                f"{getattr(s, 'quantity', 0) or 0:,.0f}",
                f"{getattr(s, 'price', 0) or 0:,.2f}",
                _fmt(s.value),
            ])
        skt = Table(sk_data, colWidths=[W*0.46, W*0.16, W*0.18, W*0.20])
        skt.setStyle(_tbl_style())
        story += [skt, Spacer(1, 3*mm)]

    # Physical assets
    cats = {
        'gold':        ('Gold', '🥇'),
        'silver':      ('Silver', '🥈'),
        'real_estate': ('Real Estate', '🏠'),
        'ppf':         ('PPF', '🏛️'),
        'vpf':         ('VPF', '🏛️'),
        'ssy':         ('Sukanya Samriddhi Yojana', '🏛️'),
        'fd':          ('Fixed Deposits', '🏦'),
        'cash':        ('Cash & Savings', '💵'),
        'other':       ('Other Assets', '📦'),
    }
    grouped = {}
    for a in assets:
        key = a.category if a.category in cats else ('real_estate' if a.category.startswith('real_estate') else 'other')
        grouped.setdefault(key, []).append(a)

    for key, (label, icon) in cats.items():
        items = grouped.get(key, [])
        if not items:
            continue
        story.append(Paragraph(f"{icon} {label}", S['h3']))
        ph_data = [['Name', 'Value (₹)']]
        for a in items:
            ph_data.append([a.name or label, _fmt(a.value)])
        ph_data.append(['Subtotal', _fmt(sum(a.value for a in items))])
        pht = Table(ph_data, colWidths=[W*0.70, W*0.30])
        pht.setStyle(_tbl_style())
        pht.setStyle(TableStyle([
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#1a1e2d')),
            ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN',      (1,0),  (1,-1),  'RIGHT'),
        ]))
        story += [pht, Spacer(1, 3*mm)]

    # ── LIFE STAGE ANALYSIS ──
    if profile and allocation:
        story += _section_header(f"🧭  Life Stage Analysis — {stage_data['icon']} {stage_data['name']}", S)
        story.append(Paragraph(stage_data['description'], S['muted']))
        story.append(Spacer(1, 3*mm))

        story.append(Paragraph(f"Portfolio Health Score: {health_score}/100", ParagraphStyle(
            'hs', fontSize=12, fontName='Helvetica-Bold',
            textColor=_GREEN if health_score >= 75 else _AMBER if health_score >= 50 else _RED,
            spaceAfter=4)))

        ls_data = [['Asset Class', 'Actual %', 'Target %', 'Gap', 'Status']]
        for b in allocation:
            gap = b['gap']
            status = '✓ On Track' if abs(gap) <= 5 else (f'▲ Under {gap}%' if gap > 0 else f'▼ Over {abs(gap)}%')
            ls_data.append([
                f"{b['icon']} {b['label']}",
                _pct(b['actual']),
                _pct(b['target']),
                f"{'+' if gap>0 else ''}{gap}%",
                status,
            ])
        lst = Table(ls_data, colWidths=[W*0.32, W*0.15, W*0.15, W*0.15, W*0.23])
        lst.setStyle(_tbl_style())
        lst.setStyle(TableStyle([('ALIGN',(1,0),(-1,-1),'CENTER')]))
        story += [lst, Spacer(1, 3*mm)]

        story.append(Paragraph("Recommendations", S['h3']))
        for rec in recommendations:
            color = _RED if rec['type'] == 'increase' else _AMBER if rec['type'] == 'reduce' else _GREEN
            story.append(Paragraph(
                f"<font color='#{color.hexval()[2:]}'>{rec['icon']} {rec['headline']}</font>",
                S['body']))
            story.append(Paragraph(rec['detail'], S['muted']))
            if rec['instruments']:
                story.append(Paragraph(
                    "  Suggested: " + " · ".join(rec['instruments']), S['muted']))
            story.append(Spacer(1, 3*mm))

    # ── GOALS + SIP PROJECTIONS ──
    if goals:
        story += _section_header("🎯  Financial Goals & SIP Projections", S)
        for g in goals:
            calc = calculate_goal(g.target_amt, g.target_year,
                                  g.current_savings, g.monthly_sip, g.annual_return)
            status_color = _GREEN if calc['on_track'] else _RED
            story.append(KeepTogether([
                Paragraph(f"{g.emoji or '⭐'} {g.name}", S['h3']),
                Spacer(1, 1*mm),
            ]))
            g_meta = [
                ['Target Amount', _fmt(g.target_amt), 'Target Year', str(g.target_year)],
                ['Current Savings', _fmt(g.current_savings or 0),
                 'Monthly SIP', _fmt(g.monthly_sip or 0)],
                ['Annual Return', f"{g.annual_return}% p.a.",
                 'Projected Corpus', _fmt(calc['projected'])],
                ['Years Left', str(calc['years_left']),
                 'Status',
                 'On Track ✓' if calc['on_track'] else f"Shortfall ₹{calc.get('shortfall',0):,.0f}"],
            ]
            gmt = Table(g_meta, colWidths=[W*0.22, W*0.28, W*0.22, W*0.28])
            gmt.setStyle(TableStyle([
                ('BACKGROUND',   (0,0), (-1,-1), _CARD),
                ('TEXTCOLOR',    (0,0), (-1,-1), _WHITE),
                ('TEXTCOLOR',    (0,0), (0,-1),  _MUTED),
                ('TEXTCOLOR',    (2,0), (2,-1),  _MUTED),
                ('FONTNAME',     (1,0), (1,-1),  'Helvetica-Bold'),
                ('FONTNAME',     (3,0), (3,-1),  'Helvetica-Bold'),
                ('FONTSIZE',     (0,0), (-1,-1),  8),
                ('GRID',         (0,0), (-1,-1),  0.3, _BORDER),
                ('LEFTPADDING',  (0,0), (-1,-1),  6),
                ('TOPPADDING',   (0,0), (-1,-1),  5),
                ('BOTTOMPADDING',(0,0), (-1,-1),  5),
            ]))
            story += [gmt, Spacer(1, 2*mm)]

            # SIP projection table
            proj = _sip_projection(g.target_amt, g.target_year,
                                   g.current_savings, g.monthly_sip, g.annual_return)
            if proj:
                story.append(Paragraph("SIP Growth Projection", S['h3']))
                sip_hdr = [['Period', 'Projected Balance (₹)', 'vs Target (₹)', 'Progress %']]
                sip_rows = []
                for (label, bal, kind) in proj:
                    delta = bal - g.target_amt
                    progress = min(round(bal / g.target_amt * 100, 1), 100) if g.target_amt else 0
                    sip_rows.append([
                        label,
                        f"{bal:,.0f}",
                        f"{'+' if delta >= 0 else ''}{delta:,.0f}",
                        f"{progress}%",
                    ])
                sip_data = sip_hdr + sip_rows
                sipt = Table(sip_data, colWidths=[W*0.22, W*0.28, W*0.28, W*0.22])
                sipt.setStyle(_tbl_style())
                sipt.setStyle(TableStyle([('ALIGN',(1,0),(-1,-1),'RIGHT')]))
                story += [sipt, Spacer(1, 5*mm)]

    # ── FOOTER ──
    story += [
        HRFlowable(width='100%', thickness=0.5, color=_BORDER),
        Spacer(1, 2*mm),
        Paragraph("Generated by MyWealthLens · For personal use only · Not investment advice",
                  ParagraphStyle('footer', fontSize=7, textColor=_MUTED,
                                 fontName='Helvetica', alignment=TA_CENTER)),
    ]

    # ── page background ──
    def _dark_bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(_DARK)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=_dark_bg, onLaterPages=_dark_bg)
    buf.seek(0)
    fname = f"MyWealthLens_{current_user.name.replace(' ','_')}_{_dt.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=fname)


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
@app.route('/export/excel')
@login_required
def export_excel():
    assets  = Asset.query.filter_by(user_id=current_user.id).all()
    mfs     = MutualFund.query.filter_by(user_id=current_user.id).all()
    stocks  = Stock.query.filter_by(user_id=current_user.id).all()
    goals   = Goal.query.filter_by(user_id=current_user.id).order_by(Goal.target_year).all()
    profile = UserProfile.query.filter_by(user_id=current_user.id).first()

    equity_val     = sum(m.value for m in mfs) + sum(s.value for s in stocks)
    debt_val       = sum(a.value for a in assets if a.category in ['ppf','vpf','ssy','fd'])
    gold_val       = sum(a.value for a in assets if a.category in ['gold','silver'])
    realestate_val = sum(a.value for a in assets if a.category.startswith('real_estate'))
    cash_val       = sum(a.value for a in assets if a.category == 'cash')
    other_val      = sum(a.value for a in assets if a.category == 'other')
    total          = equity_val + debt_val + gold_val + realestate_val + cash_val + other_val

    stage_data = allocation = recommendations = health_score = None
    if profile:
        stage_key  = detect_life_stage(profile.age, profile.marital_status, profile.dependents)
        stage_data = LIFE_STAGES[stage_key]
        _, allocation, recommendations, _, health_score, _ = \
            compute_life_stage_data(profile, assets, mfs, stocks)

    wb = openpyxl.Workbook()

    # ── colour helpers ──
    TEAL_HEX  = '00D4AA'
    DARK_HEX  = '0D0F14'
    CARD_HEX  = '13161F'
    CARD2_HEX = '1A1E2D'
    BORD_HEX  = '1E2130'
    WHITE_HEX = 'E2E8F0'
    MUTED_HEX = '64748B'
    GREEN_HEX = '22C55E'
    RED_HEX   = 'EF4444'
    AMBER_HEX = 'F59E0B'

    def _fill(hex_): return PatternFill('solid', fgColor=hex_)
    def _font(hex_=WHITE_HEX, bold=False, sz=10):
        return Font(color=hex_, bold=bold, size=sz, name='Calibri')
    def _border():
        s = Side(style='thin', color=BORD_HEX)
        return Border(left=s, right=s, top=s, bottom=s)
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _hdr_row(ws, row_num, values, col_start=1):
        for i, v in enumerate(values):
            c = ws.cell(row=row_num, column=col_start+i, value=v)
            c.fill = _fill(CARD_HEX)
            c.font = _font(TEAL_HEX, bold=True, sz=9)
            c.border = _border()
            c.alignment = _align('center')

    def _data_row(ws, row_num, values, col_start=1, alt=False):
        bg = CARD2_HEX if alt else CARD_HEX
        for i, v in enumerate(values):
            c = ws.cell(row=row_num, column=col_start+i, value=v)
            c.fill = _fill(bg)
            c.font = _font(sz=9)
            c.border = _border()
            c.alignment = _align()

    def _title_cell(ws, row_num, col, text, sz=14):
        c = ws.cell(row=row_num, column=col, value=text)
        c.font = Font(color=TEAL_HEX, bold=True, size=sz, name='Calibri')
        c.fill = _fill(DARK_HEX)
        c.alignment = _align()

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _freeze(ws, cell='A2'):
        ws.freeze_panes = cell

    def _tab_color(ws, hex_):
        ws.sheet_properties.tabColor = hex_

    # ════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    _tab_color(ws1, TEAL_HEX)
    ws1.sheet_view.showGridLines = False
    for row in ws1.iter_rows(min_row=1, max_row=60, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws1, 1, 1, "MyWealthLens — Wealth Report", sz=16)
    ws1.cell(row=2, column=1, value=f"Generated: {_dt.now().strftime('%d %B %Y, %I:%M %p')}").font = _font(MUTED_HEX, sz=9)

    r = 4
    _title_cell(ws1, r, 1, "USER PROFILE", sz=11)
    r += 1
    _hdr_row(ws1, r, ['Field', 'Value'])
    r += 1
    profile_rows = [
        ('Name', current_user.name),
        ('Email', current_user.email),
    ]
    if profile:
        profile_rows += [
            ('Age', profile.age),
            ('Life Stage', f"{stage_data['icon']} {stage_data['name']}" if stage_data else '—'),
            ('Marital Status', profile.marital_status.title()),
            ('Dependents', profile.dependents),
            ('Portfolio Health Score', f"{health_score}/100" if health_score is not None else '—'),
        ]
    for i, (k, v) in enumerate(profile_rows):
        _data_row(ws1, r, [k, v], alt=i%2==1)
        r += 1

    r += 1
    _title_cell(ws1, r, 1, "NET WORTH SUMMARY", sz=11)
    r += 1
    _hdr_row(ws1, r, ['Asset Class', 'Value (₹)', 'Allocation %'])
    r += 1
    summary_rows = [
        ('Equity (MF + Stocks)', equity_val),
        ('Debt (PPF/VPF/SSY/FD)', debt_val),
        ('Gold / Silver', gold_val),
        ('Real Estate', realestate_val),
        ('Cash & Others', cash_val + other_val),
    ]
    for i, (lbl, val) in enumerate(summary_rows):
        pct_val = round(val/total*100, 1) if total else 0
        _data_row(ws1, r, [lbl, val, f"{pct_val}%"], alt=i%2==1)
        ws1.cell(row=r, column=2).number_format = '₹#,##0'
        r += 1
    # Total row
    for col, val in enumerate([('TOTAL', total, '100%')], 1):
        pass
    _hdr_row(ws1, r, ['TOTAL', total, '100%'])
    ws1.cell(row=r, column=2).number_format = '₹#,##0'

    _set_col_widths(ws1, [28, 20, 15])
    _freeze(ws1, 'A5')

    # ════════════════════════════════════════════════
    # SHEET 2 — All Assets
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Assets")
    _tab_color(ws2, '6366F1')
    ws2.sheet_view.showGridLines = False
    for row in ws2.iter_rows(min_row=1, max_row=500, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws2, 1, 1, "All Assets", sz=14)
    r2 = 3
    _hdr_row(ws2, r2, ['Category', 'Name / Scheme', 'Sub-type', 'Units / Qty', 'Price / NAV (₹)', 'Value (₹)'])
    r2 += 1

    all_rows = []
    for m in mfs:
        all_rows.append(['Mutual Fund', m.scheme_name or m.name or '—',
                         getattr(m, 'amc', '') or '',
                         getattr(m, 'units', '') or '', getattr(m, 'nav', '') or '', m.value])
    for s in stocks:
        all_rows.append(['Stock', s.name or '—', getattr(s, 'isin', '') or '',
                         getattr(s, 'quantity', '') or '', getattr(s, 'price', '') or '', s.value])
    for a in assets:
        all_rows.append([a.category.replace('_',' ').title(), a.name or '—', '', '', '', a.value])

    for i, row_data in enumerate(all_rows):
        _data_row(ws2, r2, row_data, alt=i%2==1)
        ws2.cell(row=r2, column=6).number_format = '₹#,##0'
        r2 += 1

    _set_col_widths(ws2, [16, 40, 20, 12, 15, 16])
    _freeze(ws2, 'A4')

    # ════════════════════════════════════════════════
    # SHEET 3 — Life Stage
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet("Life Stage")
    _tab_color(ws3, '10B981')
    ws3.sheet_view.showGridLines = False
    for row in ws3.iter_rows(min_row=1, max_row=60, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws3, 1, 1, "Life Stage Analysis", sz=14)
    r3 = 3

    if profile and allocation and stage_data:
        ws3.cell(row=r3, column=1, value=f"Life Stage: {stage_data['icon']} {stage_data['name']}").font = _font(TEAL_HEX, bold=True)
        r3 += 1
        ws3.cell(row=r3, column=1, value=stage_data['description']).font = _font(MUTED_HEX, sz=9)
        r3 += 1
        ws3.cell(row=r3, column=1, value=f"Portfolio Health Score: {health_score}/100").font = \
            _font(GREEN_HEX if health_score >= 75 else AMBER_HEX if health_score >= 50 else RED_HEX, bold=True)
        r3 += 2

        _hdr_row(ws3, r3, ['Asset Class', 'Actual %', 'Target %', 'Gap', 'Status'])
        r3 += 1
        for i, b in enumerate(allocation):
            gap = b['gap']
            status = '✓ On Track' if abs(gap) <= 5 else (f'▲ Under {gap}%' if gap > 0 else f'▼ Over {abs(gap)}%')
            _data_row(ws3, r3, [f"{b['icon']} {b['label']}", f"{b['actual']}%",
                                 f"{b['target']}%", f"{'+' if gap>0 else ''}{gap}%", status], alt=i%2==1)
            # colour gap cell
            gap_color = RED_HEX if gap > 5 else AMBER_HEX if gap < -5 else GREEN_HEX
            ws3.cell(row=r3, column=4).font = _font(gap_color, bold=True, sz=9)
            r3 += 1

        r3 += 1
        _title_cell(ws3, r3, 1, "Recommendations", sz=11)
        r3 += 1
        _hdr_row(ws3, r3, ['Asset Class', 'Type', 'Action Required', 'Suggested Instruments'])
        r3 += 1
        for i, rec in enumerate(recommendations):
            insts = ' · '.join(rec.get('instruments', []))
            _data_row(ws3, r3, [rec['icon'], rec['type'].title(),
                                 rec['headline'], insts], alt=i%2==1)
            r3 += 1
    else:
        ws3.cell(row=r3, column=1, value="Life stage profile not configured.").font = _font(MUTED_HEX)

    _set_col_widths(ws3, [20, 12, 12, 10, 16, 50])
    _freeze(ws3, 'A4')

    # ════════════════════════════════════════════════
    # SHEET 4 — Goals + SIP Projections
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet("Goals & Projections")
    _tab_color(ws4, 'F59E0B')
    ws4.sheet_view.showGridLines = False
    for row in ws4.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(DARK_HEX)

    _title_cell(ws4, 1, 1, "Goals & SIP Projections", sz=14)
    r4 = 3

    if goals:
        for g in goals:
            calc = calculate_goal(g.target_amt, g.target_year,
                                  g.current_savings, g.monthly_sip, g.annual_return)
            _title_cell(ws4, r4, 1, f"{g.emoji or '⭐'} {g.name}", sz=12)
            r4 += 1
            _hdr_row(ws4, r4, ['Target (₹)', 'Target Year', 'Savings (₹)',
                                'SIP/mo (₹)', 'Return %', 'Projected (₹)', 'Status'])
            r4 += 1
            status = 'On Track ✓' if calc['on_track'] else f"Shortfall ₹{calc.get('shortfall',0):,.0f}"
            _data_row(ws4, r4, [g.target_amt, g.target_year, g.current_savings or 0,
                                 g.monthly_sip or 0, f"{g.annual_return}%",
                                 round(calc['projected']), status])
            for col in [1, 3, 4, 6]:
                ws4.cell(row=r4, column=col).number_format = '₹#,##0'
            ws4.cell(row=r4, column=7).font = \
                _font(GREEN_HEX if calc['on_track'] else RED_HEX, bold=True, sz=9)
            r4 += 2

            # SIP projection
            ws4.cell(row=r4, column=1, value="SIP Growth Projection").font = _font(TEAL_HEX, bold=True, sz=10)
            r4 += 1
            _hdr_row(ws4, r4, ['Period', 'Projected Balance (₹)', 'vs Target (₹)', 'Progress %'])
            r4 += 1

            proj = _sip_projection(g.target_amt, g.target_year,
                                   g.current_savings, g.monthly_sip, g.annual_return)
            for i, (label, bal, _) in enumerate(proj):
                delta = bal - g.target_amt
                progress = min(round(bal / g.target_amt * 100, 1), 100) if g.target_amt else 0
                _data_row(ws4, r4, [label, round(bal), round(delta), f"{progress}%"], alt=i%2==1)
                ws4.cell(row=r4, column=2).number_format = '₹#,##0'
                ws4.cell(row=r4, column=3).number_format = '₹#,##0'
                delta_color = GREEN_HEX if delta >= 0 else RED_HEX
                ws4.cell(row=r4, column=3).font = _font(delta_color, sz=9)
                r4 += 1
            r4 += 2
    else:
        ws4.cell(row=r4, column=1, value="No goals configured.").font = _font(MUTED_HEX)

    _set_col_widths(ws4, [16, 20, 18, 16])
    _freeze(ws4, 'A3')

    # ── save + send ──
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"MyWealthLens_{current_user.name.replace(' ','_')}_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)
